#!/usr/bin/env python3
"""
Delta-Import fuer "Wählerinnenverzeichnis Wien vollständig_2026.xlsx".

Workflow:
1) Neue Excel-Datei einlesen (A=Titel, B=VN, C=NN, D=Adresse)
2) Gegen bestehende DB-Adressen (PostgREST API) adressbasiert vergleichen
3) Nur neue Adressen geocodieren (Nominatim, mit Rate-Limit)
4) SQL-Delta-Import erzeugen (ohne TRUNCATE)

Hinweis:
- Das Skript fuehrt standardmaessig KEIN SQL auf der Datenbank aus.
- Es erzeugt eine SQL-Datei, die anschliessend mit psql ausgefuehrt wird.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import random
import re
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from dataclasses import dataclass, asdict

import openpyxl
import requests


DEFAULT_API = "https://204.168.217.211.nip.io/api"
DEFAULT_XLSX = "Wählerinnenverzeichnis Wien vollständig_2026.xlsx"
DEFAULT_USER_AGENT = "NOVUM-ZIV-Import-2026/1.0"
DELAY_S = 1.7
AMBIG_GEO_PREFIX = "AMBIG|"
STREET_TYPO_CORRECTIONS = {
    "strabe": "strasse",
    "strase": "strasse",
    "str": "strasse",
    "gassee": "gasse",
    "hauptsrasse": "hauptstrasse",
    "haptstrasse": "hauptstrasse",
}


@dataclass
class ParsedRow:
    row_num: int
    titel: str
    vorname: str
    nachname: str
    arzt_name: str
    raw_adresse: str
    strasse: str
    hausnummer: str
    zusatz: str
    plz: str


@dataclass
class GeocodedRow:
    parsed: ParsedRow
    lat: float | None
    lon: float | None
    geo_name: str | None
    geocode_error: str | None = None


def parse_env_file(path: str) -> None:
    if not path:
        return
    if not os.path.exists(path):
        raise FileNotFoundError(f"Env-Datei nicht gefunden: {path}")
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def normalize_text(value: str) -> str:
    value = (value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("ß", "ss")
    value = re.sub(r"\s+", " ", value)
    return value


def normalize_hnr(value: str) -> str:
    value = normalize_text(value)
    value = value.replace(" ", "")
    return value


def normalize_street_lookup(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"[^a-z0-9 ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def addr_key(plz: str, strasse: str, hausnummer: str, zusatz: str) -> tuple[str, str, str, str]:
    return (
        normalize_text(plz),
        normalize_text(strasse),
        normalize_hnr(hausnummer),
        normalize_hnr(zusatz),
    )


def split_first_last_name(full_name: str) -> tuple[str, str]:
    parts = [p for p in normalize_text(full_name).split(" ") if p]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[-1]


def name_street_key(vorname: str, nachname: str, strasse: str) -> tuple[str, str, str]:
    return (
        normalize_text(vorname),
        normalize_text(nachname),
        normalize_text(strasse),
    )


def clean_address_raw(raw: str) -> str:
    s = (raw or "").replace("\n", " ").replace("\r", " ")
    # Telefon-/Klammerteile entfernen, da sie Parsing und Geocoding stoeren.
    s = re.sub(r"\[[^\]]*\]", "", s)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def split_strasse_hnr(addr_left: str) -> tuple[str, str]:
    text = (addr_left or "").strip()
    m = re.match(r"^(.+?)\s+(\d\S*)$", text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    parts = text.rsplit(" ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return text, ""


def split_hnr_zusatz(hnr: str) -> tuple[str, str]:
    if not hnr:
        return "", ""
    # Beispiel: 31-33/2/7 -> hausnummer=31-33, zusatz=2/7
    parts = hnr.split("/", 1)
    hausnummer = parts[0].strip()
    zusatz = parts[1].strip() if len(parts) > 1 else ""
    return hausnummer, zusatz


def parse_address(raw: str) -> tuple[str, str, str, str]:
    cleaned = clean_address_raw(raw)
    if "," in cleaned:
        left, right = cleaned.split(",", 1)
    else:
        left, right = cleaned, ""
    left = left.strip()
    right = right.strip()

    plz_match = re.search(r"\b(1\d{3})\b", right)
    plz = plz_match.group(1) if plz_match else ""

    strasse, hnr_raw = split_strasse_hnr(left)
    hausnummer, zusatz = split_hnr_zusatz(hnr_raw)

    # Falls keine Hausnummer erkannt wurde, verwende die erste gefundene Zahl aus der Adresse.
    if not hausnummer:
        first_num = extract_first_house_number(cleaned)
        if first_num:
            hausnummer = first_num
            strasse = re.sub(rf"\b{re.escape(first_num)}\b", " ", strasse, count=1)
            strasse = re.sub(r"\s+", " ", strasse).strip(" ,-/")

    return strasse, hausnummer, zusatz, plz


def excel_headers(row: tuple[object, ...]) -> list[str]:
    headers = []
    for cell in row:
        if cell is None:
            headers.append("")
        else:
            headers.append(str(cell).strip())
    return headers


def find_col(headers: list[str], candidates: tuple[str, ...]) -> int:
    normalized = [normalize_text(h) for h in headers]
    for cand in candidates:
        cand_n = normalize_text(cand)
        for i, h in enumerate(normalized):
            if h == cand_n:
                return i
    raise ValueError(f"Spalte nicht gefunden. Erwartet eine von: {', '.join(candidates)}")


def parse_excel(path: str) -> list[ParsedRow]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows)
    headers = excel_headers(header_row)

    idx_titel = find_col(headers, ("Titel", "Title"))
    idx_vn = find_col(headers, ("VN", "Vorname", "First Name"))
    idx_nn = find_col(headers, ("NN", "Nachname", "Surname", "Last Name"))
    idx_addr = find_col(headers, ("Adresse", "Address"))

    parsed: list[ParsedRow] = []
    for i, row in enumerate(rows, start=2):
        titel = str(row[idx_titel] or "").strip()
        vorname = str(row[idx_vn] or "").strip()
        nachname = str(row[idx_nn] or "").strip()
        raw_addr = str(row[idx_addr] or "").strip()

        if not raw_addr or (not vorname and not nachname and not titel):
            continue

        strasse, hausnummer, zusatz, plz = parse_address(raw_addr)
        arzt_name = " ".join(p for p in [vorname, nachname] if p).strip()

        parsed.append(
            ParsedRow(
                row_num=i,
                titel=titel,
                vorname=vorname,
                nachname=nachname,
                arzt_name=arzt_name,
                raw_adresse=raw_addr,
                strasse=strasse,
                hausnummer=hausnummer,
                zusatz=zusatz,
                plz=plz,
            )
        )
    return parsed


def login(api: str, email: str, password: str, tls_verify: bool) -> str:
    r = requests.post(
        f"{api}/rpc/login",
        json={"email": email, "passwort": password},
        timeout=20,
        verify=tls_verify,
    )
    r.raise_for_status()
    payload = r.json()
    token = payload.get("token")
    if not token:
        raise RuntimeError("Login erfolgreich, aber kein token im Response.")
    return token


def fetch_existing_addresses(api: str, token: str, tls_verify: bool) -> list[dict]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    out: list[dict] = []
    limit = 1000
    offset = 0
    while True:
        query = (
            "select=id,plz,strasse,hausnummer,zusatz,arzt_name,import_batch"
            f"&order=id.asc&limit={limit}&offset={offset}"
        )
        r = requests.get(f"{api}/adressen?{query}", headers=headers, timeout=30, verify=tls_verify)
        r.raise_for_status()
        chunk = r.json()
        if not chunk:
            break
        out.extend(chunk)
        if len(chunk) < limit:
            break
        offset += limit
    return out


def build_street_plz_index(existing: list[dict]) -> dict[str, set[str]]:
    idx: dict[str, set[str]] = {}
    for row in existing:
        street = normalize_street_lookup(row.get("strasse") or "")
        plz = normalize_text(row.get("plz") or "")
        if not street or not plz:
            continue
        idx.setdefault(street, set()).add(plz)
    return idx


def enrich_missing_plz_from_existing(rows: list[ParsedRow], street_plz_idx: dict[str, set[str]]) -> int:
    enriched = 0
    for row in rows:
        if normalize_text(row.plz):
            continue
        street_key = normalize_street_lookup(row.strasse)
        plz_set = street_plz_idx.get(street_key) or set()
        if len(plz_set) == 1:
            row.plz = next(iter(plz_set))
            enriched += 1
    return enriched


def geocode(plz: str, strasse: str, hausnummer: str, user_agent: str) -> tuple[float | None, float | None, str | None, str | None]:
    q_addr = " ".join(p for p in [strasse, hausnummer] if p).strip()
    params_dict = {
        "format": "json",
        "limit": 5,
        "countrycodes": "at",
        "addressdetails": 0,
    }
    if plz:
        params_dict.update(
            {
                "street": q_addr,
                "postalcode": plz,
                "city": "Wien",
                "country": "Oesterreich",
            }
        )
    else:
        # Bei fehlender PLZ explizit auf Wien begrenzen.
        params_dict.update(
            {
                "street": q_addr,
                "city": "Wien",
                "country": "Oesterreich",
                "viewbox": "16.182,48.322,16.577,48.114",
                "bounded": 1,
            }
        )
    params = urllib.parse.urlencode(params_dict)
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": user_agent})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if not data:
            return None, None, None, "not_found"
        first = data[0]
        geo_name = (first.get("display_name") or "").split(",")[0].strip() or None
        if len(data) > 1 and geo_name:
            geo_name = AMBIG_GEO_PREFIX + geo_name
        return float(first["lat"]), float(first["lon"]), geo_name, None
    except Exception as e:
        return None, None, None, str(e)


def simplify_hnr_for_retry(hnr: str) -> str:
    """Vereinfacht komplexe Hausnummern fuer einen zweiten Geocoding-Versuch."""
    if not hnr:
        return ""
    s = hnr.strip()
    # 31-33 -> 31, 31/2 -> 31
    s = re.split(r"[-/]", s, maxsplit=1)[0].strip()
    # Falls noch Mischformat, ersten numerischen Kern verwenden.
    m = re.match(r"^(\d+[a-zA-Z]?)", s)
    return m.group(1) if m else s


def apply_street_typos(strasse: str) -> str:
    out = (strasse or "").strip()
    for wrong, right in STREET_TYPO_CORRECTIONS.items():
        out = re.sub(rf"(?i)\b{re.escape(wrong)}\b", right, out)
    return out


def extract_first_house_number(text: str) -> str:
    m = re.search(r"\b(\d+[a-zA-Z]?)\b", text or "")
    return m.group(1) if m else ""


def normalize_for_not_found_retry(strasse: str, hausnummer: str) -> tuple[str, str]:
    fixed_street = apply_street_typos(strasse)
    first_num = extract_first_house_number(fixed_street)
    if first_num and (not hausnummer or not re.search(r"\d", hausnummer)):
        hausnummer = first_num
    if first_num:
        fixed_street = re.sub(r"\b\d+[a-zA-Z]?\b", " ", fixed_street, count=1)
        fixed_street = re.sub(r"\s+", " ", fixed_street).strip(" ,-/")
    return fixed_street.strip(), hausnummer.strip()


def sql_str(s: str | None) -> str:
    if s is None or s == "":
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_text(s: str | None) -> str:
    """SQL-quoted text that preserves empty string as '' instead of NULL."""
    value = "" if s is None else str(s)
    return "'" + value.replace("'", "''") + "'"


def build_sql(rows: list[GeocodedRow], batch_name: str) -> str:
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    value_lines = []
    for g in rows:
        p = g.parsed
        value_lines.append(
            "(" + ", ".join(
                [
                    sql_text(p.plz),
                    sql_text("Wien"),
                    sql_text(p.strasse),
                    sql_text(p.hausnummer),
                    sql_str(p.zusatz or None),
                    str(g.lat),
                    str(g.lon),
                    sql_str(p.titel or None),
                    sql_str(p.arzt_name),
                    sql_str(g.geo_name),
                    sql_str(batch_name),
                ]
            ) + ")"
        )

    sql = [
        f"-- Delta-Import 2026, generiert am {ts}",
        f"-- Batch: {batch_name}",
        "BEGIN;",
        "",
        "WITH incoming(plz, ort, strasse, hausnummer, zusatz, lat, lon, titel, arzt_name, geo_name, import_batch) AS (",
        "  VALUES",
        "  " + ",\n  ".join(value_lines),
        "),",
        "inserted AS (",
        "  INSERT INTO adressen (id, plz, ort, strasse, hausnummer, zusatz, lat, lon, titel, arzt_name, geo_name, import_batch)",
        "  SELECT",
        "    uuid_generate_v4(), i.plz, i.ort, i.strasse, i.hausnummer, i.zusatz, i.lat, i.lon,",
        "    i.titel, i.arzt_name, i.geo_name, i.import_batch",
        "  FROM incoming i",
        "  WHERE NOT EXISTS (",
        "    SELECT 1",
        "    FROM adressen a",
        "    WHERE lower(trim(a.plz)) = lower(trim(i.plz))",
        "      AND lower(regexp_replace(trim(a.strasse), '\\s+', ' ', 'g')) = lower(regexp_replace(trim(i.strasse), '\\s+', ' ', 'g'))",
        "      AND lower(replace(trim(a.hausnummer), ' ', '')) = lower(replace(trim(i.hausnummer), ' ', ''))",
        "      AND lower(replace(trim(coalesce(a.zusatz, '')), ' ', '')) = lower(replace(trim(coalesce(i.zusatz, '')), ' ', ''))",
        "  )",
        "  RETURNING id",
        ")",
        "SELECT",
        "  (SELECT COUNT(*) FROM incoming) AS incoming_rows,",
        "  (SELECT COUNT(*) FROM inserted) AS inserted_rows;",
        "",
        "COMMIT;",
    ]
    return "\n".join(sql) + "\n"


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Delta-Import aus Wählerinnenverzeichnis 2026")
    p.add_argument("--xlsx", default=DEFAULT_XLSX, help="Pfad zur Excel-Datei")
    p.add_argument("--api", default=os.getenv("NOVUMZIV_API", DEFAULT_API), help="PostgREST API URL")
    p.add_argument("--email", default=os.getenv("NOVUMZIV_EMAIL"), help="Login E-Mail")
    p.add_argument("--password", default=os.getenv("NOVUMZIV_PASS"), help="Login Passwort")
    p.add_argument("--tls-verify", action="store_true", default=os.getenv("NOVUMZIV_TLS_VERIFY", "1") == "1")
    p.add_argument("--no-tls-verify", action="store_false", dest="tls_verify")
    p.add_argument("--env-file", default="", help="Optionale env-Datei (KEY=VALUE)")
    p.add_argument("--batch", default=f"import_wien_2026_{dt.datetime.now().strftime('%Y%m%d')}", help="Import-Batch Name")
    p.add_argument("--skip-geocode", action="store_true", help="Geocoding ueberspringen (nur Delta-Zaehlung)")
    p.add_argument("--out-sql", default="import_delta_2026.sql", help="Ausgabe SQL-Datei")
    p.add_argument("--out-new-json", default="import_delta_2026_new.json", help="Ausgabe nur neue Datensaetze")
    p.add_argument(
        "--out-geocoded-json",
        default="import_delta_2026_geocoded.json",
        help="Checkpoint-Datei fuer erfolgreich geocodierte Delta-Datensaetze",
    )
    p.add_argument("--out-failed-json", default="import_delta_2026_geocode_failed.json", help="Ausgabe geocode-fehlgeschlagen")
    p.add_argument("--user-agent", default=DEFAULT_USER_AGENT, help="User-Agent fuer Nominatim")
    p.add_argument("--geocode-delay", type=float, default=DELAY_S, help="Basis-Delay zwischen Geocoding-Requests")
    p.add_argument("--geocode-max-retries", type=int, default=6, help="Maximale Retry-Versuche bei 429")
    p.add_argument("--geocode-backoff-base", type=float, default=4.0, help="Backoff-Basis in Sekunden bei 429")
    p.add_argument(
        "--compare-mode",
        choices=("address", "name_street"),
        default="address",
        help="Vergleichslogik: 'address' = PLZ+Strasse+Hausnummer+Zusatz, 'name_street' = Vorname+Nachname+Strasse",
    )
    return p


def main() -> int:
    args = build_parser().parse_args()

    try:
        parse_env_file(args.env_file)

        if not args.email:
            args.email = os.getenv("NOVUMZIV_EMAIL")
        if not args.password:
            args.password = os.getenv("NOVUMZIV_PASS")

        if not args.email or not args.password:
            raise SystemExit("NOVUMZIV_EMAIL und NOVUMZIV_PASS fehlen. Setze env vars oder nutze --email/--password.")

        print(f"[1/6] Lese Excel: {args.xlsx}")
        parsed = parse_excel(args.xlsx)
        if not parsed:
            raise SystemExit("Keine importierbaren Zeilen in Excel gefunden.")
        print(f"      Importierbare Zeilen: {len(parsed)}")

        print("[2/6] Login API")
        token = login(args.api, args.email, args.password, args.tls_verify)

        print("[3/6] Lade bestehenden DB-Bestand")
        existing = fetch_existing_addresses(args.api, token, args.tls_verify)
        print(f"      Bestehende DB-Adressen: {len(existing)}")
        street_plz_idx = build_street_plz_index(existing)

        if args.compare_mode == "address":
            existing_keys = {
                addr_key(
                    e.get("plz") or "",
                    e.get("strasse") or "",
                    e.get("hausnummer") or "",
                    e.get("zusatz") or "",
                )
                for e in existing
            }
        else:
            existing_keys = set()
            for e in existing:
                first, last = split_first_last_name(e.get("arzt_name") or "")
                existing_keys.add(
                    name_street_key(first, last, e.get("strasse") or "")
                )

        if args.compare_mode == "address":
            print("[4/6] Delta-Vergleich (adressbasiert)")
        else:
            print("[4/6] Delta-Vergleich (Vorname+Nachname+Strasse)")
        new_rows: list[ParsedRow] = []
        seen_new_keys: set[tuple[str, ...]] = set()
        duplicates_in_excel = 0

        for p in parsed:
            if args.compare_mode == "address":
                key = addr_key(p.plz, p.strasse, p.hausnummer, p.zusatz)
            else:
                key = name_street_key(p.vorname, p.nachname, p.strasse)
            if key in existing_keys:
                continue
            if key in seen_new_keys:
                duplicates_in_excel += 1
                continue
            seen_new_keys.add(key)
            new_rows.append(p)

        print(f"      Neue Adressen (Delta): {len(new_rows)}")
        if duplicates_in_excel:
            print(f"      Doppelte Zeilen innerhalb Excel (ignoriert): {duplicates_in_excel}")

        plz_enriched = enrich_missing_plz_from_existing(new_rows, street_plz_idx)
        if plz_enriched:
            print(f"      PLZ via Strassenmuster nachangereichert: {plz_enriched}")

        with open(args.out_new_json, "w", encoding="utf-8") as f:
            json.dump([asdict(r) for r in new_rows], f, ensure_ascii=False, indent=2)
        print(f"      Delta-JSON geschrieben: {args.out_new_json}")

        if not new_rows:
            print("[5/6] Keine neuen Adressen gefunden. Kein SQL erzeugt.")
            return 0

        geocoded: list[GeocodedRow] = []
        failed: list[GeocodedRow] = []

        if args.skip_geocode:
            print("[5/6] Geocoding uebersprungen (--skip-geocode)")
        else:
            print("[5/6] Geocoding fuer Delta-Adressen")
            checkpoint: dict[str, dict] = {}
            if os.path.exists(args.out_geocoded_json):
                with open(args.out_geocoded_json, "r", encoding="utf-8") as f:
                    prior = json.load(f)
                for item in prior:
                    key = f"{item.get('row_num')}|{normalize_text(item.get('arzt_name') or '')}|{normalize_text(item.get('strasse') or '')}"
                    checkpoint[key] = item
                if checkpoint:
                    print(f"      Resume-Checkpoint geladen: {len(checkpoint)}")

            checkpoint_list: list[dict] = list(checkpoint.values())
            for i, p in enumerate(new_rows, start=1):
                p_key = f"{p.row_num}|{normalize_text(p.arzt_name)}|{normalize_text(p.strasse)}"

                if p_key in checkpoint:
                    cp = checkpoint[p_key]
                    lat = cp.get("lat")
                    lon = cp.get("lon")
                    geo_name = cp.get("geo_name")
                    err = cp.get("geocode_error")
                    row = GeocodedRow(parsed=p, lat=lat, lon=lon, geo_name=geo_name, geocode_error=err)
                else:
                    current_street = p.strasse
                    current_hnr = p.hausnummer
                    lat = lon = geo_name = None
                    err: str | None = None
                    attempt = 0
                    normalized_retry_done = False

                    while attempt <= args.geocode_max_retries:
                        lat, lon, geo_name, err = geocode(p.plz, current_street, current_hnr, args.user_agent)
                        if lat is not None and lon is not None:
                            break

                        is_429 = err is not None and "429" in err
                        if is_429 and attempt < args.geocode_max_retries:
                            backoff = min(90.0, args.geocode_backoff_base * (2 ** attempt))
                            backoff += random.uniform(0.0, 1.0)
                            time.sleep(backoff)
                            attempt += 1
                            continue

                        # Einmalige Normalisierungsschicht bei not_found: Tippfehler + Nummernextraktion.
                        if err == "not_found" and not normalized_retry_done:
                            retry_street, retry_hnr = normalize_for_not_found_retry(current_street, current_hnr)
                            if retry_street != current_street or retry_hnr != current_hnr:
                                current_street = retry_street
                                current_hnr = retry_hnr
                                normalized_retry_done = True
                                attempt += 1
                                continue

                        # Einmaliger Fallback mit vereinfachter Hausnummer bei not_found.
                        if (
                            err == "not_found"
                            and attempt == 0
                            and current_hnr
                            and current_hnr != simplify_hnr_for_retry(current_hnr)
                        ):
                            current_hnr = simplify_hnr_for_retry(current_hnr)
                            attempt += 1
                            continue

                        break

                    row = GeocodedRow(parsed=p, lat=lat, lon=lon, geo_name=geo_name, geocode_error=err)

                    if lat is not None and lon is not None:
                        checkpoint_item = {
                            "row_num": p.row_num,
                            "arzt_name": p.arzt_name,
                            "strasse": p.strasse,
                            "hausnummer": p.hausnummer,
                            "plz": p.plz,
                            "lat": lat,
                            "lon": lon,
                            "geo_name": geo_name,
                            "geocode_error": None,
                        }
                        checkpoint[p_key] = checkpoint_item
                        checkpoint_list = list(checkpoint.values())
                        if len(checkpoint_list) % 10 == 0:
                            with open(args.out_geocoded_json, "w", encoding="utf-8") as f:
                                json.dump(checkpoint_list, f, ensure_ascii=False, indent=2)

                if lat is not None and lon is not None:
                    geocoded.append(row)
                else:
                    failed.append(row)

                if i <= 3 or i % 25 == 0:
                    status = "OK" if lat is not None else "FAIL"
                    print(f"      [{i}/{len(new_rows)}] {status}: {p.strasse} {p.hausnummer}, {p.plz}")
                if p_key not in checkpoint or (lat is None or lon is None):
                    time.sleep(max(0.0, args.geocode_delay))

            with open(args.out_geocoded_json, "w", encoding="utf-8") as f:
                json.dump(checkpoint_list, f, ensure_ascii=False, indent=2)
            print(f"      Geocode-Checkpoint: {args.out_geocoded_json}")

            with open(args.out_failed_json, "w", encoding="utf-8") as f:
                json.dump([asdict(r) for r in failed], f, ensure_ascii=False, indent=2)

            print(f"      Geocoding erfolgreich: {len(geocoded)}")
            print(f"      Geocoding fehlgeschlagen: {len(failed)}")
            print(f"      Fehlerliste: {args.out_failed_json}")

        print("[6/6] SQL-Datei erzeugen")
        if args.skip_geocode:
            if os.path.exists(args.out_geocoded_json):
                with open(args.out_geocoded_json, "r", encoding="utf-8") as f:
                    cp_rows = json.load(f)

                geocoded_from_cp: list[GeocodedRow] = []
                parsed_by_row = {r.row_num: r for r in new_rows}
                for c in cp_rows:
                    row_num = int(c.get("row_num") or 0)
                    parsed_row = parsed_by_row.get(row_num)
                    if not parsed_row:
                        continue
                    lat = c.get("lat")
                    lon = c.get("lon")
                    if lat is None or lon is None:
                        continue
                    geocoded_from_cp.append(
                        GeocodedRow(
                            parsed=parsed_row,
                            lat=float(lat),
                            lon=float(lon),
                            geo_name=c.get("geo_name"),
                            geocode_error=None,
                        )
                    )

                if geocoded_from_cp:
                    sql = build_sql(geocoded_from_cp, args.batch)
                    with open(args.out_sql, "w", encoding="utf-8") as f:
                        f.write(sql)
                    print(f"      SQL aus Checkpoint geschrieben: {args.out_sql}")
                    print(f"      Verwendete Checkpoint-Datensaetze: {len(geocoded_from_cp)}")
                    return 0

            print("      SQL-Export uebersprungen, da ohne Koordinaten kein Insert moeglich ist.")
            return 0

        if not geocoded:
            print("      Keine geocodierten Delta-Adressen vorhanden. Kein SQL geschrieben.")
            return 0

        sql = build_sql(geocoded, args.batch)
        with open(args.out_sql, "w", encoding="utf-8") as f:
            f.write(sql)

        print(f"      SQL geschrieben: {args.out_sql}")
        print("\nFertig.")
        print(f"- Rohzeilen (Excel): {len(parsed)}")
        print(f"- Delta (neu): {len(new_rows)}")
        print(f"- Geocodiert (insertierbar): {len(geocoded)}")
        print(f"- Geocode-Fehler: {len(failed)}")
        print("\nNaechster Schritt:")
        print(f"psql -d <DB_NAME> -f {args.out_sql}")
        return 0

    except requests.HTTPError as e:
        print(f"HTTP-Fehler: {e}", file=sys.stderr)
        if e.response is not None:
            print(e.response.text, file=sys.stderr)
        return 2
    except Exception as e:
        print(f"Fehler: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
